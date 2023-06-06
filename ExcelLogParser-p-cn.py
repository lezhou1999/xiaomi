"""
功能描述：
此脚本主要用来解析自动下载、解压和分析wifi log

使用方法:
1. 安装python3.0或以上版本, 将python的安装目录下的Scripts路径加到系统的环境变量里面 .
   https://www.python.org/downloads/release/python-360/
2. 安装python第三方库：xlrd, xlwt, xlutils, urllib, codecs, zipfile, tkinter, shutil，re命令行下运行 easy_install xlrd ...... 来安装以上第三方库.
3. 将附件里面的ExcelLogParser.py 保存到一个新建文件夹中.
4. 从feedback系统导出反馈excel，记得勾上日志，下载到本地并保存为Excel 97-2003格式 文件(.xls），由于第三方库的局限只能解析这种格式的文件.
5. 进入步骤3目录命令行下运行： ExcelLogParser.py   xxxx.xls  等待运行完成。（建议把 ExcelLogParser.py 和xxx.xls文件保存到同一个新建的目录下否则要加路径)
6. 步骤5运行时会在.xls文件目录下新建文件夹保存所有自动下载的对应反馈问题的log
7. 运行完成后会生成一个result.txt保存所有的问题分析和找到的关键log，并且生成一个excel文件来统计每个反馈对应的问题统计

如何增加关键字：
1. 所有关键字添加请按照顺序加载列表的尾部
2. 如果要增加的关键字在bugreport里面，可以加在keywords里面
3. 如果要增加的关键字在wlan driver log里面，可以加在driver_keywords里面
4. 如果要增加的关键字在wlan fw log里面(目前只支持qcom)，可以加在firmware_keywords里面
5. 加入关键字后，就会在results里面看到相关log的分析结果，但excel不会有变化
6. 如果要在excel做些统计：
   a) 在issueStat增加相关字段，构造函数也增加对应的初始化的值
   b) 并在代码里有new这个对象的地方的参数都做响应修改
   c) 还要再write_sheet的data也加上增加的字段
   d) 最后需要在bugreport、driver以及fw log的解析函数中，根据keywords排放的位置判断，填充issueStat的新增字段：find_file_text、find_driver_file_text、find_firmware_file_text
   d) 以上增加请按照参数的顺序对齐
"""
import sys
import re
import os
import shutil
import zipfile
import traceback
import urllib.request
import xlrd
import xlwt
from xlutils.copy import copy
import cchardet as chardet
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment

#from LogParser import *  # 本地同级目录下的LogParser模块
from numba import jit

keywords_id = {}
keywords_count = {}
feadback_id = ''
feadback_title = ''
feadback_sw_ver = ''
feadback_issue_type = ''
fdb_id_info = {}

User_Feedback_Is_855 = True  # 是855平台的用户反馈excel:true  不是:false
keywords = [
    'freq:(.*), protocol:.*, .*router:(.*), totalTC',
    'default network switch to slave wifi close all sockets',
    'NetworkDiagnostics.*TYPE=G',
    'NetworkDiagnostics:.*ret=false',
    'NetworkDiagnostics:.*FAILED',
    'NetworkDiagnostics: Network diagnostics start',
    ' NetworkDiagnostics:',
    'time=.*event=abnormaltrx',
    'mStaticIpConfig: IP address',
    # 'NUD_FAILED',
    # 'PROBE_.*FAIL',
    # 'handleNetworkUnvalidated',
    # 'handleNetworkNoInternet',
    # 'delay: 460',
    # 'wifiLink:fail',
    # 'netSet:fail',
    # 'signal:fail',
    # 'topo:fail',
    # 'dns:fail',
    # 'Conn:fail',
    # 'latency:fail',
]
driver_keywords = [
    # 'enable_dynamic_nss = 1',
    'data stall: abnormaltrx=DIR:TX,Event:Hang',
    'data stall: abnormaltrx=DIR:TX,event:AbDrop',
    'data stall: abnormaltrx=DIR:TXRX,event:BTWifiCoexLow',
    'data stall: abnormaltrx=DIR:TX,event:AbRate',
    'data stall: abnormaltrx',
    'halIsTxHang:.*timeout\[sec:([0-9]*)\]'
]
firmware_keywords = ['consecutive failure=(.*), kickout thresh=512',
                     'peerid_tidnum_hwq:(.*) ppdu_id.*seqnum__sifs__sch:(0x[a-zA-Z0-9]*) tries__fes__flush:'
                     ]
# table_name = 'bugreport'
hwq_no_ack_count = {}
hwq_no_ack_count_dict = {}

suffix = ['txt']


class issueStat(object):
    def __init__(self, id, freq, router, nss, tx_fail, tx_no_ack, dual_wifi_issue, ndg, ndr,
                 ndf, nds, ndt, dst, txhangt, dsa, dsc, dsaa, dso, staticIp, sw_ver, type, desc):
        self.feadback_Id = id
        self.diag_freq = freq
        self.diag_router = router
        self.dynamic_nss = nss
        self.con_tx_fail = tx_fail
        self.tx_no_ack = tx_no_ack
        self.dual_wifi_issue = dual_wifi_issue
        self.netdiag_gateway = ndg
        self.netdiag_retfalse = ndr
        self.netdiag_failed = ndf
        self.netdiag_start = nds
        self.netdiag_trigger = ndt
        self.datastall_txhang = dst
        self.txHangTimeout = txhangt
        self.datastall_abdrop = dsa
        self.datastall_coexlow = dsc
        self.datastall_abrate = dsaa
        self.datastall_others = dso
        self.staticip = staticIp
        self.sw_ver = sw_ver
        self.issue_type = type
        self.issue_desc = desc


def get_encoding(fi):
    with open(fi, "rb") as f:
        data = f.read()
        result = chardet.detect(data)
        return result['encoding']


def find_file_text(root_dir, target_text, issue_stat):
    result = ''
    global feadback_id, feadback_title, feadback_sw_ver, feadback_issue_type
    for root, dirs, files in os.walk(root_dir):
        # result += root+'\n'
        for file in files:
            # file_suffix = file[file.find('.') + 1:len(file)]
            file_suffix = os.path.splitext(file)[-1][1:]
            ##            if (os.path.splitext(file)[0].find('bugreport-') == -1):
            ##                return
            if file_suffix in suffix:
                print(file)
                file_encoding = get_encoding(os.path.join(root, file))
                # result += '-->'+file+'\n'
                with open(os.path.join(root, file), 'r', encoding=file_encoding, errors='ignore') as f:
                    for line in f.readlines():
                        for key in target_text:
                            key_results = re.search(key, line, re.IGNORECASE)
                            if key_results != None:
                                if target_text.index(key) == 0:
                                    issue_stat.diag_freq = int(key_results.group(1))
                                    issue_stat.diag_router = key_results.group(2)
                                # elif target_text.index(key) == 1:
                                #     issue_stat.dual_wifi_issue = 'SwitchCauseDestorySocket'
                                elif target_text.index(key) == 2:
                                    issue_stat.netdiag_gateway = 'GatewayFail'
                                elif target_text.index(key) == 3:
                                    issue_stat.netdiag_retfalse = 'DiagRetFalse'
                                elif target_text.index(key) == 4:
                                    issue_stat.netdiag_failed = 'DiagFailed'
                                elif target_text.index(key) == 5:
                                    # print('----------------------')
                                    # print(line)
                                    # print(key)
                                    issue_stat.netdiag_start = 'DiagStart'
                                elif target_text.index(key) == 6:
                                    issue_stat.netdiag_trigger = 'DiagTrigger'
                                elif target_text.index(key) == 8:
                                    issue_stat.staticip = 'Static IP'
                                result += line
                                break
            elif (file_suffix == 'zip'):
                result += '-->' + file + '\n'
                result += 'zip file: need unzip it firstly'
    return result


def find_driver_file_text(root_dir, target_text, issue_stat):
    result = ''
    tx_hang_max_timeout = 0
    global feadback_id, feadback_title, feadback_sw_ver, feadback_issue_type
    for root, dirs, files in os.walk(root_dir):
        # result += root+'\n'
        for file in files:
            if os.path.splitext(file)[0].find('kernel_log_') == -1:
                continue
            # print(file)
            file_encoding = get_encoding(os.path.join(root, file))
            # result += '-->'+file+'\n'
            with open(os.path.join(root, file), 'r', encoding=file_encoding, errors='ignore') as f:
                for line in f.readlines():
                    for key in target_text:
                        key_results = re.search(key, line, re.IGNORECASE)
                        if key_results != None:
                            if target_text.index(key) == 0:
                                issue_stat.datastall_txhang = 'dataStall:TxHang'
                            elif target_text.index(key) == 1:
                                issue_stat.datastall_abdrop = 'dataStall:AbDrop'
                            elif target_text.index(key) == 2:
                                issue_stat.datastall_coexlow = 'dataStall:CoexLow'
                            elif target_text.index(key) == 3:
                                issue_stat.datastall_abrate = 'dataStall:AbRate'
                            elif target_text.index(key) == 4:
                                issue_stat.datastall_others = 'dataStall:Others'
                            elif target_text.index(key) == 5:
                                tx_hang_max_timeout = int(key_results.group(1))
                                if tx_hang_max_timeout > issue_stat.txHangTimeout:
                                    issue_stat.txHangTimeout = tx_hang_max_timeout

                            result += line
                            break
            # print(issue_stat.txHangTimeout)
    return result


def stat_no_ack(tx_fail_status, hwq_id, issue_stat):
    global hwq_no_ack_count, hwq_no_ack_count_dict
    if (tx_fail_status & 0x0000000f) != 4:
        if hwq_no_ack_count.get(hwq_id) != None:
            hwq_no_ack_count[hwq_id] = 0
        return False
    else:
        if hwq_no_ack_count.get(hwq_id) == None:
            hwq_no_ack_count[hwq_id] = 1
        else:
            hwq_no_ack_count[hwq_id] = hwq_no_ack_count[hwq_id] + 1
            if hwq_no_ack_count_dict.get(hwq_id) != None:
                if hwq_no_ack_count_dict[hwq_id] < hwq_no_ack_count[hwq_id]:
                    hwq_no_ack_count_dict[hwq_id] = hwq_no_ack_count[hwq_id]
            else:
                hwq_no_ack_count_dict[hwq_id] = hwq_no_ack_count[hwq_id]

            if hwq_no_ack_count_dict[hwq_id] > 48:
                issue_stat.tx_no_ack = hwq_no_ack_count_dict[hwq_id]
    return True


def find_firmware_file_text(root_dir, target_text, issue_stat):
    result = ''
    global feadback_id, feadback_title, feadback_sw_ver, feadback_issue_type, hwq_no_ack_count_dict, hwq_no_ack_count

    hwq_no_ack_count.clear()
    hwq_no_ack_count_dict.clear()
    for root, dirs, files in os.walk(root_dir):
        # result += root+'\n'
        for file in files:
            if os.path.splitext(file)[0].find('cnss_fw_logs_') == -1:
                continue
            # print(file)
            file_encoding = get_encoding(os.path.join(root, file))
            # result += '-->'+file+'\n'
            with open(os.path.join(root, file), 'r', encoding=file_encoding, errors='ignore') as f:
                for line in f.readlines():
                    for key in target_text:
                        key_results = re.search(key, line, re.IGNORECASE)
                        if key_results != None:
                            if target_text.index(key) == 0:
                                freq = '0000'
                                tx_fail_count = key_results.group(1)
                                if int(tx_fail_count) > 100:
                                    if issue_stat.con_tx_fail < int(tx_fail_count):
                                        issue_stat.con_tx_fail = int(tx_fail_count)
                                else:
                                    continue
                            elif target_text.index(key) == 1:
                                hwq_id = int(key_results.group(1), 16) & 0x000000ff
                                tx_fail_status = int(key_results.group(2), 16)
                                if not stat_no_ack(tx_fail_status, hwq_id, issue_stat):
                                    continue
                            result += line
                            break
    return result


def parser_one_feedback(ws, log_name):
    global User_Feedback_Is_855, feadback_id, feadback_title, feadback_sw_ver, feadback_issue_type
    global hwq_no_ack_count, hwq_no_ack_count_dict
    is_zip = 1
    bugreport_name = ''
    parser_result = ''
    try:
        if zipfile.is_zipfile(log_name):
            zip1_file_name = log_name
            file_zip1 = zipfile.ZipFile(zip1_file_name, 'r')
            bugreport_name = os.path.splitext(zip1_file_name)[0]
            try:
                os.mkdir(bugreport_name)
            except FileExistsError:
                print('文件夹已经存在，不用创建！')

            # file_zip1.extractall(bugreport_name)
            '''
            将第二级压缩文件名默认值设为第一级压缩文件文件名，
            以适配F11部分用户反馈下载的压缩文件的第一级压缩文件直接是第二级压缩文件的不规范睿智情况
            后续如果用户反馈的网站规范后，zip2_file_name = zip1_file_name可以删除
            '''

            zip2_file_name = zip1_file_name
            for file in file_zip1.namelist():
                # 如果文件长度为1，
                if len(file_zip1.namelist()) == 1:
                    file_zip1.extract(file, bugreport_name)  # 解压出这个文件
                    zip2_file_name = bugreport_name + '\\' + file
                    break
                if os.path.splitext(file)[0].find('bugreport-') != -1:
                    file_zip1.extract(file, bugreport_name)
                    zip2_file_name = bugreport_name + '\\' + file
                elif os.path.splitext(file)[0].find('host_driver_logs_') != -1:
                    file_zip1.extract(file, bugreport_name)
                elif os.path.splitext(file)[0].find('cnss_fw_logs_') != -1:
                    file_zip1.extract(file, bugreport_name)
                elif os.path.splitext(file)[0].find('kernel_log_') != -1:
                    file_zip1.extract(file, bugreport_name)
            file_zip1.close()
            if zipfile.is_zipfile(zip2_file_name):
                file_zip2 = zipfile.ZipFile(zip2_file_name, 'r')
                for file in file_zip2.namelist():
                    if (os.path.splitext(file)[0].find('bugreport-') != -1):
                        file_zip2.extract(file, bugreport_name)
                        break
                file_zip2.close()
                # os.remove(zip2_file_name)


    except:
        print('Exception when extract the zip file!')
        return ''

    if bugreport_name == '':
        return ''
    ##    parser_result = parser_log(bugreport_name)
    ##    if os.path.exists(bugreport_name) and is_zip != 0:
    ##        #删除log文件
    ##        os.remove(bugreport_name)
    issue_stat = issueStat(feadback_id, 0000, 'unknown', 'unknown', 0, 'unknown', 'unknown',
                           'N', 'N', 'N', 'N', 'N', 'N', 0, 'N', 'N', 'N', 'N', 'N',
                           feadback_sw_ver, feadback_issue_type, feadback_title)

    print('filename:' + bugreport_name)
    # print('---->start analyze bugreport\n')
    # parser_result += '---->start analyze bugreport\n'
    # parser_result += find_file_text(bugreport_name, keywords, issue_stat)

    print('---->start analyze wlan driver log\n')
    parser_result += '---->start analyze wlan driver log\n'
    parser_result += find_driver_file_text(bugreport_name, driver_keywords, issue_stat)
    #
    # print('---->start analyze firmware log')
    # parser_result += '---->start analyze wlan driver log'
    # parser_result += find_firmware_file_text(bugreport_name, firmware_keywords, issue_stat)

    write_sheet(ws, issue_stat)
    max_hwq_count = 0
    # for key in hwq_no_ack_count_dict:
    #     print('---->hwq: ' + str(key) + ' tx_no_ack_fail_count: ' + str(hwq_no_ack_count_dict[key]) + '\n')
    #     parser_result += '---->hwq: ' + str(key) + ' tx_no_ack_fail_count: ' + str(hwq_no_ack_count_dict[key]) + '\n'
    #     if hwq_no_ack_count_dict[key] > max_hwq_count:
    #         max_hwq_count = hwq_no_ack_count_dict[key]

    if (os.path.isdir(bugreport_name)):
        shutil.rmtree(bugreport_name)
    return parser_result


def write_sheet(worksheet, issue_stat):
    data = (issue_stat.feadback_Id, issue_stat.diag_freq, issue_stat.diag_router, issue_stat.dynamic_nss,
            issue_stat.con_tx_fail, issue_stat.tx_no_ack, issue_stat.dual_wifi_issue, issue_stat.netdiag_gateway,
            issue_stat.netdiag_retfalse, issue_stat.netdiag_failed, issue_stat.netdiag_start,
            issue_stat.netdiag_trigger,
            issue_stat.datastall_txhang, issue_stat.txHangTimeout,issue_stat.datastall_abdrop, issue_stat.datastall_coexlow,
            issue_stat.datastall_abrate,
            issue_stat.datastall_others, issue_stat.staticip,
            issue_stat.sw_ver, issue_stat.issue_type, issue_stat.issue_desc)
    worksheet.append(data)


def main():
    global User_Feedback_Is_855, feadback_id, feadback_title, feadback_sw_ver, feadback_issue_type
    file_type = ''
    parser_result = ''
    if len(sys.argv) <= 1:
        print('Please input a file name as an parameter!')
        file_name = '0113_0225-L11-1677302782528.xls'
        # return
    else:
        file_name = sys.argv[1]

    file_type = os.path.splitext(file_name)
    if file_type[1] == '.zip':
        parser_result = parser_one_feedback(file_name)
        print(parser_result)
        print(input('\nInput Enter to EXIT!'))
        return
    elif file_type[1] == '.log' or file_type[1] == '.txt':
        parser_result = parser_one_feedback(file_name)
        print(parser_result)
        print(input('\nInput Enter to EXIT!'))
        return
    elif file_type[1] == '.xls':
        excel_name = file_name
    else:
        print('No valid input file!')
        print(input('\nInput Enter to EXIT!'))
        return

    rdata = xlrd.open_workbook(excel_name, formatting_info=True)  # 读取excel文件
    wdata = copy(rdata)  # 浅copy
    try:
        rtable = rdata.sheet_by_name(u'Modem')  # 通过名称获取一个工作表
        wtable = wdata.get_sheet(u'Modem')  # 打开sheet
        User_Feedback_Is_855 = False
    except Exception as e:  # 855平台的用户反馈excel只要一张工作表
        print('没有名为Modem的工作表，将分析第一张工作表')
        rtable = rdata.sheet_by_index(0)  # 将分析第一张工作表
        wtable = wdata.get_sheet(0)  # 打开sheet
        User_Feedback_Is_855 = True
    nrows = rtable.nrows  # 获取行数
    ncols = rtable.ncols  # 获取列数

    if User_Feedback_Is_855:  # 855平台的用户反馈
        for colnum in range(0, ncols):
            if rtable.cell(0, colnum).value == '具体问题(用户)':
                anl_col = colnum
            if rtable.cell(0, colnum).value == '日志':
                log_col = colnum
            if rtable.cell(0, colnum).value == '反馈内容':
                fdb_col = colnum
            elif rtable.cell(0, colnum).value == '反馈ID':
                fdb_id = colnum
            elif rtable.cell(0, colnum).value == '问题发生时间':
                fdb_time = colnum
            elif rtable.cell(0, colnum).value == 'ROM版本':
                fdb_rom_ver = colnum
            # if rtable.cell(0, colnum).value == '标注师':#2081.9.27  有的excel中没有标注师这一列
            # lab_col = colnum
    else:
        for colnum in range(0, ncols):
            if rtable.cell(0, colnum).value == '具体问题':
                anl_col = colnum
            if rtable.cell(0, colnum).value == '日志文件':
                log_col = colnum
            if rtable.cell(0, colnum).value == '反馈内容':
                fdb_col = colnum
            elif rtable.cell(0, colnum).value == '反馈ID':
                fdb_id = colnum
            # if rtable.cell(0, colnum).value == '标注师':#2081.9.27  有的excel中没有标注师这一列
            # lab_col = colnum
    # 2018.8.27 设置单元格宽度
    wtable.col(anl_col).width = 256 * 60
    # wtable.col(lab_col).width=256*30
    style = xlwt.XFStyle()
    # 写入的边框属性
    borders = xlwt.Borders()
    borders.left = 1
    borders.right = 1
    borders.top = 1
    borders.bottom = 1
    borders.bottom_colour = 0x3A
    style.borders = borders
    # 设置单元格对齐方式
    alignment = xlwt.Alignment()  # 创建alignment
    # alignment.horz = xlwt.Alignment.HORZ_CENTER   #设置水平对齐为居中，May be: HORZ_GENERAL, HORZ_LEFT, HORZ_CENTER, HORZ_RIGHT, HORZ_FILLED, HORZ_JUSTIFIED, HORZ_CENTER_ACROSS_SEL, HORZ_DISTRIBUTED
    alignment.vert = xlwt.Alignment.VERT_CENTER  # 设置垂直对齐为居中，May be: VERT_TOP, VERT_CENTER, VERT_BOTTOM, VERT_JUSTIFIED, VERT_DISTRIBUTED
    style.alignment = alignment  # 应用alignment到style上

    log_folder = excel_name + '_log'
    try:
        os.mkdir(log_folder)
    except FileExistsError:
        print('文件夹已经存在，不用创建！')

    # ws = wb.create_sheet(table_name)
    # ws_title1 = ('Id', 'Title', 'url')
    # ws_title2 = tuple(keywords)
    # ws_title = ws_title1 + ws_title2
    # ws.append(ws_title)

    # dayTime = datetime.now().strftime('%Y-%m-%d')
    wb = Workbook()
    ws = wb.active
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet('feadback statis')

    ws_title = issueStat('ID', 'Freq', 'Router', 'DynamicNss', 'Consecutive tx failure', 'max tx no ack count',
                         'DualWifiDestroySocket', 'DiagGatewayFail', 'DiagRetFail', 'DiagFailed', 'DiagStart',
                         'DiagTrigger', 'DataStallTxHang', 'txHangTimeout', 'DataStallAbDrop', 'DataStallCoexLow', 'DataStallAbRate',
                         'DataStallOthers', 'Static Ip', 'SW Ver', 'Issue Type', 'Issue Title')
    write_sheet(ws, ws_title)

    with open('result.txt', 'w') as f:
        for rownum in range(1, nrows):
            # if rtable.cell(rownum, anl_col).value != '':
            # continue
            print('开始分析问题问题:', rownum + 1)
            report = '开始分析问题问题:' + str(rownum + 1) + '\n'
            report += '反馈ID: ' + str(rtable.cell(rownum, fdb_id).value) + '\n'
            report += '反馈内容: ' + rtable.cell(rownum, fdb_col).value + '\n'
            report += '具体问题(用户)：' + rtable.cell(rownum, anl_col).value + '\n'
            report += '问题发生时间：' + rtable.cell(rownum, fdb_time).value + '\n'
            report += 'ROM版本：' + rtable.cell(rownum, fdb_rom_ver).value + '\n'

            feadback_id = rtable.cell(rownum, fdb_id).value
            feadback_title = rtable.cell(rownum, fdb_col).value
            feadback_title = feadback_title.replace('\n', '')
            feadback_sw_ver = rtable.cell(rownum, fdb_rom_ver).value
            feadback_issue_type = rtable.cell(rownum, anl_col).value

            log_url = rtable.cell(rownum, log_col).value

            # ws_data = [feadback_id, rtable.cell(rownum, fdb_col).value, log_url]
            # 日志下载链接变更
            # 如果有feedbackId，则固定用http://feedback.pt.xiaomi.com/openapi/getFeedbackLogFile?feedbackId=Id号访问
            # 否则，正常访问
            if log_url.find('feedbackId') != -1:
                match = re.search(r'feedbackId=(\d+)', log_url)
                if match:
                    log_url = match.group(1)
                else:
                    print('下载链接不正确')
                log_url = "http://feedback.pt.xiaomi.com/openapi/getFeedbackLogFile?feedbackId=" + log_url
            if log_url == '' and rtable.cell(rownum, fdb_col).value == '':  # excel结尾的空行，可能有几百行空行，不要分析，白白浪费时间
                break
            elif log_url == '':
                wtable.write(rownum, anl_col, '没有log', style)
                print('没有log')
                report += '没有log\n'
                report += '----------------------------------------------------------------------------------\n'
                issue_stat = issueStat(feadback_id, 0000, '没有log', '没有log', 0, '没有log', '没有log',
                                       'N', 'N', 'N', 'N', 'N', 'N', 0, 'N', 'N', 'N', 'N', 'N',
                                       feadback_sw_ver, feadback_issue_type, feadback_title)
                write_sheet(ws, issue_stat)
                f.write(report)
                continue
            try:
                print(log_url)
                log_name = os.getcwd() + '\\' + log_folder + '\\' + str(rownum + 1) + ".zip"  # 得到标准log名称
                if os.path.exists(log_name):
                    print('日志: ' + log_name + ' 已经存在，不再重新下载....')
                else:
                    urllib.request.urlretrieve(log_url, log_name)  # 从网上下载log 存在log_name中
            #                #shutil.move(log_name, os.getcwd() + '\\' + excel_name + '\\' + log_name)
            except Exception as e:  # urllib.request.urlretrieve 调用失败
                wtable.write(rownum, anl_col, '下载log文件失败', style)
                print(traceback.format_exc())
                print('Download log failed!')
                continue
            parser_result = parser_one_feedback(ws, log_name)
            if parser_result == '':
                continue
            report += parser_result
            report += '----------------------------------------------------------------------------------\n'
            print(parser_result)  # 打印出结果

            f.write(report)
        # for cell in ws.columns:
        #     cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        output_file = "issue_stat_" + excel_name + '.xlsx'
        wb.save(output_file)
        ##        wtable.write(rownum, anl_col, parser_result,style)#向单元格写入结果
        ##        try:   #try语句缩进1格,每分析完一次结果就保存一次，中间报错使之前已经分析的结果得以保存---高策
        ##            wdata.save(excel_name) #把结果写入excel   2018.8.27改 将保存文件提出for循环 防止最后一次没有LOG直接continue跳出循环而没有保存最后一次用户反馈
        ##        except PermissionError:
        ##            print('写Ecxel失败，请确认关闭了Ecxel！')
        ##            break
        f.write(report)

    print('\n所有问题分析完成！')

    print(input('\nInput Enter to EXIT!'))


def format_result(str):
    return str.ljust(25, ' ')


if __name__ == '__main__':
    main()
